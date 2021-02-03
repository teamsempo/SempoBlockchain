import pytest, os
from faker.providers import phone_number
from faker import Faker
from server.models.user import User
from server.utils.auth import get_complete_auth_token
from server.utils.executor import get_job_key
from server import red, db
from openpyxl import Workbook
import json
import time

fake = Faker()
fake.add_provider(phone_number)


def fake_csv():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='first_name')
    ws.cell(row=1, column=2, value='last_name')
    ws.cell(row=1, column=3, value='phone')
    ws.cell(row=2, column=1, value='foo1')
    ws.cell(row=2, column=2, value='bar1')
    ws.cell(row=2, column=3, value=fake.msisdn())

    wb.save('spreadsheet.xlsx')

    my_xlsx = os.path.join("spreadsheet.xlsx")

    return my_xlsx

def delete_fake_csv():
    if os.path.exists("spreadsheet.xlsx"):
        os.remove('spreadsheet.xlsx')

@pytest.mark.parametrize("tier, status_code", [
    ("subadmin", 403),
    ("admin", 200),
])
def test_spreadsheet_upload_api(test_client, authed_sempo_admin_user, tier, status_code):
    if tier:
        authed_sempo_admin_user.set_held_role('ADMIN', tier)
        auth = get_complete_auth_token(authed_sempo_admin_user)
    else:
        auth = None

    data = dict()
    data['spreadsheet'] = (fake_csv(), 'spreadsheet.xlsx')
    response = test_client.post(
        f"/api/v1/spreadsheet/upload/",
        headers=dict(
            Authorization=auth,
            Accept='multipart/form-data',
        ),
        data=data
    )

    assert response.status_code == status_code
    delete_fake_csv()


def test_dataset_api(test_client, authed_sempo_admin_user):
    data = {
        "data": [
            {
                "0": "Alf",
                "1": "Melmac",
                "2": "19027192211"
            },
            {
                "0": "Alf",
                "1": "Tanner",
                "2": "19027192211" # Same phone number, should trigger update
            },
            {
                "0": "Willie",
                "1": "Tanner",
                "2": "19027192222"
            }
        ],
        "headerPositions": {
            "0": "first_name",
            "1": "last_name",
            "2": "phone"
        },
        "country": "",
        "saveName": "",
        "isVendor": False
    }

    auth = get_complete_auth_token(authed_sempo_admin_user)
    
    response = test_client.post(
        f"/api/v1/dataset/",
        headers=dict(
            Authorization=auth,
            Accept='application/json',
        ),
        json=data
    )
    # This is necessary because the session is now active in a different thread and pytest still needs DB access
    db.session.close()
    db.engine.dispose()

    redis_id = get_job_key(authed_sempo_admin_user.id, response.json['task_uuid']) 
    status = { 'percent_complete': 0 }
    while not status['percent_complete'] == 100.0:
        time.sleep(.01)
        redis_status = red.get(redis_id)
        if redis_status:
            status = json.loads(redis_status)
    assert status['message'] == 'success'
    assert status['percent_complete'] == 100.0
    assert status['diagnostics'] == [['User Created', 200], ['User Updated', 200], ['User Created', 200]]
