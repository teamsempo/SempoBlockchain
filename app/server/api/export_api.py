from flask import Blueprint, request, make_response, jsonify, g, current_app
from flask.views import MethodView
from openpyxl import Workbook
from datetime import datetime, timedelta
import random, string, os
from sqlalchemy import and_, or_
from sqlalchemy.orm import joinedload

from dateutil import parser

from server.schemas import transfer_account_schema

from server.models.credit_transfer import CreditTransfer
from server.models.transfer_account import TransferAccount
from server.utils.transfer_enums import TransferTypeEnum, TransferStatusEnum
from server.models.user import User
from server.utils.auth import requires_auth
from server.utils.amazon_s3 import upload_local_file_to_s3
from server.utils.date_magic import find_last_period_dates
from server.utils.amazon_ses import send_export_email
from server.utils.export import generate_pdf_export, export_workbook_via_s3

export_blueprint = Blueprint('export', __name__)


class ExportAPI(MethodView):
    @requires_auth(allowed_roles={'ADMIN': 'admin'})
    def post(self):

        post_data = request.get_json()

        export_type = post_data.get('export_type')
        include_transfers = post_data.get('include_transfers')  # True or False
        include_custom_attributes = post_data.get('include_custom_attributes')  # True or False

        user_type = post_data.get('user_type')  # Beneficiaries, Vendors, All
        selected = post_data.get('selected')
        #  TODO: implement date_range
        date_range = post_data.get('date_range')  # last day, previous week, or all

        payable_period_type   = post_data.get('payable_period_type', 'month')        # day, week, month
        payable_period_length = post_data.get('payable_period_length', 1)      # Integer, ie "payable every _2_ months"
        payable_epoch         = post_data.get('payable_epoch')              # When did the very first cycle begin?

        payable_period_start_date = post_data.get('payable_period_start_date') # any sort of reasonable date string for custom date period
        payable_period_end_date = post_data.get('payable_period_end_date')

        transfer_account_columns = [
            {'header': 'Account ID',            'query_type': 'db',     'query': 'id'},
            {'header': 'User ID',               'query_type': 'custom', 'query': 'user_id'},
            {'header': 'First Name',            'query_type': 'custom', 'query': 'first_name'},
            {'header': 'Last Name',             'query_type': 'custom', 'query': 'last_name'},
            {'header': 'Public Serial Number',  'query_type': 'custom', 'query': 'public_serial_number'},
            {'header': 'Phone',                 'query_type': 'custom', 'query': 'phone'},
            {'header': 'Created (UTC)',         'query_type': 'db',     'query': 'created'},
            {'header': 'Approved',              'query_type': 'db',     'query': 'is_approved'},
            {'header': 'Beneficiary',           'query_type': 'custom', 'query': 'has_beneficiary_role'},
            {'header': 'Vendor',                'query_type': 'custom', 'query': 'has_vendor_role'},
            {'header': 'Location',              'query_type': 'custom', 'query': 'location'},
            {'header': 'Current Balance',       'query_type': 'custom', 'query': 'balance'},
            {'header': 'Amount Received',       'query_type': 'custom', 'query': 'received'},
            {'header': 'Amount Sent',           'query_type': 'custom', 'query': 'sent'}
            # {'header': 'Prev. Period Payable',  'query_type': 'custom', 'query': 'prev_period_payable'},
            # {'header': 'Total Payable',         'query_type': 'custom', 'query': 'total_payable'},
        ]

        credit_transfer_columns = [
            {'header': 'ID',                'query_type': 'db',     'query': 'id'},
            {'header': 'Transfer Amount',   'query_type': 'custom', 'query': 'transfer_amount'},
            {'header': 'Created',           'query_type': 'db',     'query': 'created'},
            {'header': 'Resolved Date',     'query_type': 'db',     'query': 'resolved_date'},
            {'header': 'Transfer Type',     'query_type': 'enum',   'query': 'transfer_type'},
            {'header': 'Transfer Type',     'query_type': 'enum', 'query': 'transfer_subtype'},
            {'header': 'Transfer Status',   'query_type': 'enum',   'query': 'transfer_status'},
            {'header': 'Sender ID',         'query_type': 'db',     'query': 'sender_transfer_account_id'},
            {'header': 'Recipient ID',      'query_type': 'db',     'query': 'recipient_transfer_account_id'},
            {'header': 'Transfer Uses',     'query_type': 'custom', 'query': 'transfer_usages'},
        ]

        # need to add Balance (Payable)

        random_string = ''.join(random.choices(string.ascii_letters, k=5))
        # TODO MAKE THIS API AUTHED
        time = str(datetime.utcnow())

        base_filename = current_app.config['DEPLOYMENT_NAME'] + '-id' + str(g.user.id) + '-' + str(time[0:10]) + '-' + random_string
        workbook_filename = base_filename + '.xlsx'
        # e.g. dev-id1-2018-09-19-asfi.xlsx
        pdf_filename = base_filename + '.pdf'

        wb = Workbook()
        ws = wb.active
        ws.title = "transfer_accounts"
        # ws1 = wb.create_sheet(title='transfer_accounts')

        start_date = None
        end_date = None
        user_filter = None

        # Create transfer_accounts workbook headers
        for index, column in enumerate(transfer_account_columns):
            _ = ws.cell(column=index + 1, row=1, value="{0}".format(column['header']))

        user_accounts = []
        credit_transfer_list = []

        # filter user accounts
        if user_type == 'beneficiary':
            user_filter = User.has_beneficiary_role

        if user_type == 'vendor':
            user_filter = User.has_vendor_role

        if date_range == 'all':
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(weeks=520)

        if date_range == 'day':
            # return previous day of transactions
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(days=1)

        if date_range == 'week':
            # return previous week of transactions
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(weeks=1)

        if user_filter is not None:
            user_accounts = User.query.filter(
                user_filter==True
            ).options(
                joinedload(User.transfer_accounts)
            ).all()

        elif user_type == 'selected':
            transfer_accounts = TransferAccount.query.filter(TransferAccount.id.in_(selected)).all()
            user_accounts = [ta.primary_user for ta in transfer_accounts]

        else:
            user_accounts = User.query.filter(
                or_(User.has_beneficiary_role == True, User.has_vendor_role == True)
            ).options(
                joinedload(User.transfer_accounts)
            ).all()

        if export_type == 'pdf':
            file_url = generate_pdf_export(user_accounts, pdf_filename)
            response_object = {
                'message': 'Export file created.',
                'data': {
                    'file_url': file_url,
                }
            }

            return make_response(jsonify(response_object)), 201

        if user_accounts is not None:

            custom_attribute_columns = []

            for index, user_account in enumerate(user_accounts):
                transfer_account = user_account.transfer_account

                if transfer_account:
                    for jindix, column in enumerate(transfer_account_columns):
                        if column['query_type'] == 'db':
                            cell_contents = "{0}".format(getattr(transfer_account, column['query']))

                        elif column['query'] == 'user_id':

                            cell_contents = "{0}".format(transfer_account.primary_user.id)

                        elif column['query'] == 'first_name':

                            cell_contents = "{0}".format(transfer_account.primary_user.first_name)

                        elif column['query'] == 'last_name':

                            cell_contents = "{0}".format(transfer_account.primary_user.last_name)

                        elif column['query'] == 'phone':

                            cell_contents = "{0}".format(transfer_account.primary_user.phone or '')

                        elif column['query'] == 'public_serial_number':

                            cell_contents = "{0}".format(transfer_account.primary_user.public_serial_number or '')

                        elif column['query'] == 'location':

                            cell_contents = "{0}".format(transfer_account.primary_user._location)

                        elif column['query'] == 'balance':
                            cell_contents = getattr(transfer_account, column['query'])/100

                        elif column['query'] == 'has_beneficiary_role':
                            cell_contents = "{0}".format(transfer_account.primary_user.has_beneficiary_role)

                        elif column['query'] == 'has_vendor_role':
                            cell_contents = "{0}".format(transfer_account.primary_user.has_vendor_role)

                        elif column['query'] == 'received':
                            received_amount = transfer_account.total_received
                            cell_contents = received_amount / 100

                        elif column['query'] == 'sent':
                            sent_amount = transfer_account.total_sent
                            cell_contents = sent_amount / 100

                        elif column['query'] == 'prev_period_payable':

                            if payable_period_start_date and payable_period_end_date:
                                prior_period_start = parser.parse(payable_period_start_date)
                                prior_period_end   = parser.parse(payable_period_end_date)
                            else:

                                if payable_epoch is None:
                                    payable_epoch = transfer_account.created
                                else:
                                    payable_epoch = parser.parse(str(payable_epoch))

                                prior_period_start, prior_period_end = find_last_period_dates(payable_epoch, datetime.utcnow(),payable_period_type,payable_period_length)

                            in_for_period = 0

                            # All funds start of last period, up to end of last period, NOT including after last period
                            in_transactions = CreditTransfer.query.filter(and_(
                                CreditTransfer.recipient_transfer_account_id == transfer_account.id,
                                CreditTransfer.transfer_status == TransferStatusEnum.COMPLETE,
                                CreditTransfer.resolved_date.between(prior_period_start,prior_period_end)
                            )).all()

                            for transaction in in_transactions:
                                in_for_period += transaction.transfer_amount

                            out_for_period = 0

                            # Out transactions DO NOT include reimbursements (withdrawals) from the previous month
                            out_transactions = CreditTransfer.query.filter(and_(
                                CreditTransfer.sender_transfer_account_id == transfer_account.id,
                                CreditTransfer.transfer_status == TransferStatusEnum.COMPLETE,
                                CreditTransfer.resolved_date.between(prior_period_start, prior_period_end),
                                CreditTransfer.transfer_type != TransferTypeEnum.WITHDRAWAL
                            )).all()

                            for transaction in out_transactions:
                                out_for_period += transaction.transfer_amount

                            payable_balance = in_for_period - out_for_period

                            cell_contents = payable_balance/100

                        elif column['query'] == 'total_payable':

                            if  payable_period_end_date:
                                prior_period_end   = parser.parse(payable_period_end_date)
                            else:

                                if payable_epoch is None:
                                    payable_epoch = transfer_account.created
                                else:
                                    payable_epoch = parser.parse(str(payable_epoch))

                                prior_period_start, prior_period_end = find_last_period_dates(payable_epoch, datetime.utcnow(),payable_period_type,payable_period_length)

                            in_for_period = 0

                            # All funds in from epoch, up to end of last period, NOT including after last period
                            in_transactions = CreditTransfer.query.filter(and_(
                                CreditTransfer.recipient_transfer_account_id == transfer_account.id,
                                CreditTransfer.transfer_status == TransferStatusEnum.COMPLETE,
                                CreditTransfer.resolved_date.between(payable_epoch,prior_period_end)
                            )).all()

                            for transaction in in_transactions:
                                in_for_period += transaction.transfer_amount

                            out_for_period = 0

                            # All funds out over all time, _including_ withdrawals
                            out_transactions = CreditTransfer.query.filter(and_(
                                CreditTransfer.sender_transfer_account_id == transfer_account.id,
                                CreditTransfer.transfer_status == TransferStatusEnum.COMPLETE,
                                CreditTransfer.resolved_date.between(payable_epoch, datetime.utcnow())
                            )).all()

                            for transaction in out_transactions:
                                out_for_period += transaction.transfer_amount

                            payable_balance = in_for_period - out_for_period

                            cell_contents = payable_balance/100

                        else:
                            cell_contents = ""

                        _ = ws.cell(column=jindix + 1, row=index + 2, value=cell_contents)

                    if include_custom_attributes:
                        # Add custom attributes as columns at the end
                        for attribute in transfer_account.primary_user.custom_attributes:
                            try:
                                col_num = custom_attribute_columns.index(attribute.name) + 1 + len(transfer_account_columns)
                            except ValueError:
                                custom_attribute_columns.append(attribute.name)
                                col_num = len(custom_attribute_columns) + len(transfer_account_columns)

                            _ = ws.cell(column=col_num, row=index + 2, value=attribute.value)

                else:
                    print('No Transfer Account for user account id: ', user_account.id)

            # Add custom attribute headers:
            if include_custom_attributes:
                for index, column_name in enumerate(custom_attribute_columns):
                    _ = ws.cell(
                        column=index + 1 + len(transfer_account_columns),
                        row=1,
                        value="{0}".format(column_name)
                    )

        if include_transfers and user_accounts is not None:
            base_credit_transfer_query = CreditTransfer.query.enable_eagerloads(False).order_by(CreditTransfer.id)

            if selected and user_filter is None:
                base_credit_transfer_query = base_credit_transfer_query.filter(
                    or_(CreditTransfer.sender_transfer_account_id.in_(selected),
                        CreditTransfer.recipient_transfer_account_id.in_(selected)))

            if start_date and end_date is not None:
                credit_transfer_list = base_credit_transfer_query.filter(
                    CreditTransfer.created.between(start_date, end_date)
                ).all()

            if date_range == 'all':
                credit_transfer_list = base_credit_transfer_query.all()

            transfer_sheet = wb.create_sheet(title='credit_transfers')

            # Create credit_transfers workbook headers
            for index, column in enumerate(credit_transfer_columns):
                _ = transfer_sheet.cell(column=index + 1, row=1, value="{0}".format(column['header']))

            if credit_transfer_list is not None:
                for index, credit_transfer in enumerate(credit_transfer_list):
                    for jindix, column in enumerate(credit_transfer_columns):
                        if column['query_type'] == 'db':
                            cell_contents = "{0}".format(getattr(credit_transfer, column['query']))
                        elif column['query_type'] == 'enum':
                            enum = getattr(credit_transfer, column['query'])
                            cell_contents = "{0}".format(enum and enum.value)
                        elif column['query'] == 'transfer_amount':
                            cell_contents = "{0}".format(getattr(credit_transfer, column['query'])/100)
                        elif column['query'] == 'transfer_usages':
                            cell_contents = ', '.join([usage.name for usage in credit_transfer.transfer_usages])
                        else:
                            cell_contents = ""

                        _ = transfer_sheet.cell(column=jindix + 1, row=index + 2, value=cell_contents)


            else:
                print('No Credit Transfers')

        if len(user_accounts) is not 0:
            file_url = export_workbook_via_s3(wb, workbook_filename)

            response_object = {
                'message': 'Export file created.',
                'data': {
                    'file_url': file_url,
                }
            }

            return make_response(jsonify(response_object)), 201

        else:
            # return no data
            response_object = {
                'message': 'No data available for export',
            }

            return make_response(jsonify(response_object)), 404


class MeExportAPI(MethodView):
    @requires_auth
    def post(self):
        post_data = request.get_json()

        transfer_account = g.user.default_transfer_account or g.user.transfer_accounts[0]

        date_range = post_data.get('date_range')  # last day, previous week, or all
        email = post_data.get('email')

        credit_transfer_columns = [
            {'header': 'ID',                'query_type': 'db',     'query': 'id'},
            {'header': 'Transfer Amount',   'query_type': 'custom', 'query': 'transfer_amount'},
            {'header': 'Created',           'query_type': 'db',     'query': 'created'},
            {'header': 'Resolved Date',     'query_type': 'db',     'query': 'resolved_date'},
            {'header': 'Transfer Type',     'query_type': 'enum', 'query': 'transfer_type'},
            {'header': 'Transfer Status',   'query_type': 'enum', 'query': 'transfer_status'},
            {'header': 'Transfer Uses',     'query_type': 'custom',  'query': 'transfer_usages'},
        ]

        random_string = ''.join(random.choices(string.ascii_letters, k=5))
        export_time = str(datetime.utcnow())
        workbook_filename = 'vendor-id' + str(transfer_account.id) + '-' + str(export_time[0:10]) + '-' + random_string + '.xlsx'
        # e.g. vendor-id9-2018-09-19-adsff.xlsx

        # Vendor Export
        wb = Workbook()
        ws = wb.active
        ws.title = "credit_transfers"

        start_date = None
        end_date = None
        credit_transfer_list = []

        # Create credit_transfers workbook headers
        for index, column in enumerate(credit_transfer_columns):
            _ = ws.cell(column=index + 1, row=1, value="{0}".format(column['header']))

        if date_range == 'day':
            # return previous day of transactions
            start_date = datetime.utcnow()
            end_date = start_date - timedelta(days=1)

        if date_range == 'week':
            # return previous week of transactions
            start_date = datetime.utcnow()
            end_date = start_date - timedelta(weeks=1)

        if start_date and end_date is not None:
            # filter by date_range & transfer_account.id
            credit_transfer_list = CreditTransfer.query.filter(
                and_(CreditTransfer.created.between(start_date, end_date), (
                or_(CreditTransfer.recipient_transfer_account_id == transfer_account.id,
                    CreditTransfer.sender_transfer_account_id == transfer_account.id))))\
                    .enable_eagerloads(False)

        else:
            # default to all credit transfers of transfer_account.id
            credit_transfer_list = CreditTransfer.query.filter(
                or_(CreditTransfer.recipient_transfer_account_id == transfer_account.id,
                    CreditTransfer.sender_transfer_account_id == transfer_account.id))\
                    .enable_eagerloads(False)

        # loop over all credit transfers, create cells
        if credit_transfer_list is not None:
            for index, credit_transfer in enumerate(credit_transfer_list):
                for jindix, column in enumerate(credit_transfer_columns):
                    if column['query_type'] == 'db':
                        cell_contents = "{0}".format(getattr(credit_transfer, column['query']))
                    elif column['query_type'] == 'enum':
                        enum = getattr(credit_transfer, column['query'])
                        cell_contents = "{0}".format(enum and enum.value)
                    elif column['query'] == 'transfer_amount':
                        cell_contents = "{0}".format(getattr(credit_transfer, column['query']) / 100)
                    elif  column['query'] == 'transfer_usages':
                        cell_contents = ', '.join([usage.name for usage in credit_transfer.transfer_usages])
                    else:
                        cell_contents = ""

                    _ = ws.cell(column=jindix + 1, row=index + 2, value=cell_contents)

        if credit_transfer_list is not None:
            file_url = export_workbook_via_s3(wb, workbook_filename, email)

            response_object = {
                'status': 'success',
                'message': 'Export file created.',
                'file_url': file_url,
            }

            return make_response(jsonify(response_object)), 201

        else:
            # return no data
            response_object = {
                'status': 'Fail',
                'message': 'No data available for export',
                'file_url': '',
            }

            return make_response(jsonify(response_object)), 404


# add Rules for API Endpoints
export_blueprint.add_url_rule(
    '/export/',
    view_func=ExportAPI.as_view('export_view'),
    methods=['POST']
)

export_blueprint.add_url_rule(
    '/me/export/',
    view_func=MeExportAPI.as_view('me_export_view'),
    methods=['POST']
)
